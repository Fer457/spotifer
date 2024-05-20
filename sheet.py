import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict

def create_or_open_excel(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Spotify Log"
        sheet.append(["Artist", "Track", "Start Time", "End Time", "Duration (minutes)"])
    return workbook, sheet

def log_track_to_excel(sheet, artist, track, start_time, end_time, duration_minutes):
    sheet.append([artist, track, str(start_time), str(end_time), duration_minutes])

def convert_to_minutes(duration_str):
    h, m, s = map(float, duration_str.split(':'))
    return h * 60 + m + s / 60

def get_daily_summary(sheet):
    artist_time = defaultdict(float)
    track_time = defaultdict(float)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        artist, track, start_time, end_time, duration_str = row
        duration_minutes = convert_to_minutes(duration_str)
        artist_time[artist] += duration_minutes
        track_time[track] += duration_minutes
    return artist_time, track_time

def save_excel(workbook, file_name):
    workbook.save(file_name)
